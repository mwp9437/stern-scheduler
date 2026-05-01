"""
Fall 2026 catalog ingest.

Reads the Stern admin xlsx and upserts one row per Class Nbr into
public.courses, collapsing multi-pattern source rows into a single row
with a JSONB `meeting_patterns` array.

Usage:
    # Direct upsert (requires SUPABASE_SERVICE_ROLE_KEY in env):
    python fall_2026.py

    # Dry-run (parse + summarize, no writes):
    python fall_2026.py --dry-run

    # Emit SQL instead of writing:
    python fall_2026.py --sql-out fall_2026_data.sql

    # Override xlsx path:
    python fall_2026.py --xlsx /path/to/file.xlsx

For Spring 2027: copy this script, swap the source path/term, sanity-check
the SESSION_TO_DURATION map for any new codes, and run.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import openpyxl


# ---- Mappings ----------------------------------------------------------

# Stern Session code → courses.duration_type. Source of truth — date math
# is intentionally absent. Unknown codes raise so future ingests don't
# silently misclassify.
SESSION_TO_DURATION: dict[str, str] = {
    "DFT": "Full Semester",
    "EFT": "Full Semester",
    "DM1": "First Half",
    "EM1": "First Half",
    "S":   "First Half",   # 9-week multi-pattern Aug 17 – Oct 20
    "DM2": "Second Half",
    "EM2": "Second Half",
    "INS": "Intensive",
    "INW": "Intensive",
    "XTL": "DBi",          # Global immersion (J-term)
}

# Source uses single-letter S/U for Sat/Sun; app expects two-letter Sa/Su.
DAY_LETTER_REMAP: dict[str, str] = {"S": "Sa", "U": "Su"}

# Range patterns that appear in INS/XTL rows. Expand to explicit day list.
RANGE_PAT_EXPANSION: dict[str, str] = {
    "M-U":  "MTWRFSaSu",   # Mon through Sun
    "M-S":  "MTWRFSa",     # Mon through Sat
    "M-FU": "MTWRFSu",     # Mon-Fri + Sun
    "M-F":  "MTWRF",       # Mon through Fri (just in case)
}

# Order matters for the canonical letter sequence (after remap).
DAY_NAMES = {
    "M": "Mon", "T": "Tue", "W": "Wed", "R": "Thu", "F": "Fri",
    "Sa": "Sat", "Su": "Sun",
}


# ---- Pat (meeting day pattern) normalization --------------------------

def normalize_pat(raw: str) -> str:
    """Return Pat with S→Sa, U→Su, and range patterns expanded.

    Examples:
        "MW"    -> "MW"
        "TR"    -> "TR"
        "S"     -> "Sa"
        "U"     -> "Su"
        "M-U"   -> "MTWRFSaSu"
        "FSU"   -> "FSaSu"
    """
    if not raw:
        return ""
    if raw in RANGE_PAT_EXPANSION:
        return RANGE_PAT_EXPANSION[raw]

    out: list[str] = []
    i = 0
    while i < len(raw):
        ch = raw[i]
        if ch in DAY_LETTER_REMAP:
            out.append(DAY_LETTER_REMAP[ch])
        elif ch in {"M", "T", "W", "R", "F"}:
            out.append(ch)
        else:
            # Unknown char (whitespace, etc.) — drop with a warning.
            print(f"  WARN: unknown day-letter {ch!r} in Pat={raw!r}", file=sys.stderr)
        i += 1
    return "".join(out)


def expand_meeting_days(normalized_pat: str) -> list[str]:
    """Split normalized pat into day tokens, e.g. 'MTWRFSaSu' -> ['M','T','W','R','F','Sa','Su']."""
    days: list[str] = []
    i = 0
    while i < len(normalized_pat):
        if i + 1 < len(normalized_pat) and normalized_pat[i:i+2] in {"Sa", "Su"}:
            days.append(normalized_pat[i:i+2])
            i += 2
        else:
            days.append(normalized_pat[i])
            i += 1
    return days


# ---- Time / date formatting -------------------------------------------

def format_time_24h(t: time | None) -> str | None:
    """time(13, 30) -> '13:30:00'. None passes through."""
    if t is None:
        return None
    return t.strftime("%H:%M:%S")


def format_time_12h_lower(t: time | None) -> str:
    """time(13, 30) -> '1:30 pm'. Empty string for None."""
    if t is None:
        return ""
    hour = t.hour
    suffix = "am" if hour < 12 else "pm"
    h12 = hour % 12 or 12
    return f"{h12}:{t.minute:02d} {suffix}"


def format_dates_full(start: datetime | date | None, end: datetime | date | None) -> str:
    """Format as 'MM/DD-MM/DD' to match the Spring no-year convention."""
    if not start or not end:
        return ""
    return f"{start.month:02d}/{start.day:02d}-{end.month:02d}/{end.day:02d}"


def format_meeting_times_full(meeting_days: str, start: time | None, end: time | None) -> str | None:
    """'MW 1:30 pm - 2:50 pm' or None if no times."""
    if start is None or end is None:
        return None
    return f"{meeting_days} {format_time_12h_lower(start)} - {format_time_12h_lower(end)}"


# ---- Pattern + course building ----------------------------------------

def build_pattern(row: dict[str, Any]) -> dict[str, Any]:
    """Convert one xlsx row into a meeting_patterns entry."""
    raw_pat = (row["Pat"] or "").strip()
    pat = normalize_pat(raw_pat)
    mtg_start: time | None = row["Mtg Start"]
    mtg_end: time | None = row["Mtg End"]
    is_async = mtg_start is None or mtg_end is None
    return {
        "pat_nbr": int(row["Pat Nbr"]),
        "meeting_days": pat,
        "start_time": format_time_24h(mtg_start),
        "end_time": format_time_24h(mtg_end),
        "dates_full": format_dates_full(row["Mtg Start Date"], row["Mtg End Date"]),
        "is_async": is_async,
    }


def _pattern_span_days(p: dict[str, Any]) -> int:
    """Days covered by a pattern's date range, or 0 if dates can't be parsed."""
    df = p.get("dates_full") or ""
    m = re.match(r"^(\d{2})/(\d{2})-(\d{2})/(\d{2})$", df)
    if not m:
        return 0
    sm, sd, em, ed = (int(x) for x in m.groups())
    # Approx: assume same year; rolls over for Dec→Jan but we won't see that
    # within Fall (Sep-Dec) or Spring (Jan-May) ranges.
    start = date(2026, sm, sd) if sm >= 6 else date(2027, sm, sd)
    end = date(2026, em, ed) if em >= sm else date(2027, em, ed)
    return (end - start).days + 1


def pick_primary(patterns: list[dict[str, Any]]) -> dict[str, Any]:
    """Longest in-person pattern (by date span). Ties: lowest pat_nbr.
    Falls back to longest async pattern if no in-person ones exist.

    Rationale: the calendar should show a class's *dominant* meeting time,
    not a 2-week intro block or a one-day make-up session. For Stats CN 2738
    (3 patterns: MWF Aug 17-28 + TR Sep 3-Oct 15 + T Oct 20), this picks
    TR Sep 3-Oct 15.
    """
    in_person = [p for p in patterns if not p["is_async"]]
    pool = in_person if in_person else patterns
    return max(pool, key=lambda p: (_pattern_span_days(p), -p["pat_nbr"]))


def describe_secondary_pattern(p: dict[str, Any]) -> str:
    """Human-readable line for the notes field."""
    days = p["meeting_days"] or "(no days)"
    if p["is_async"]:
        # Async block: just date range.
        return f"Async block: {p['dates_full']} ({days})"
    # In-person: days + time + dates.
    s = format_time_12h_lower(time.fromisoformat(p["start_time"])) if p["start_time"] else "?"
    e = format_time_12h_lower(time.fromisoformat(p["end_time"])) if p["end_time"] else "?"
    return f"Also meets: {days} {s} - {e} ({p['dates_full']})"


def build_notes(rows: list[dict[str, Any]], primary_pat_nbr: int, mode: str | None) -> str | None:
    """Concatenate: secondary Descr + non-default mode + secondary pattern summaries."""
    parts: list[str] = []

    # Secondary Descr (col 23 = 'Descr_2' in our row dicts) — same value
    # repeats across the multi-row class, so dedupe.
    descr2_values = {(r.get("Descr_2") or "").strip() for r in rows if r.get("Descr_2")}
    for v in sorted(d for d in descr2_values if d):
        parts.append(v)

    # Mode flag for non-in-person classes.
    if mode and mode != "P":
        mode_label = {"OL": "Online (live)", "OB": "Online (asynchronous)"}.get(mode, f"Mode: {mode}")
        parts.append(mode_label)

    # Secondary patterns.
    patterns = sorted([build_pattern(r) for r in rows], key=lambda p: p["pat_nbr"])
    for p in patterns:
        if p["pat_nbr"] == primary_pat_nbr:
            continue
        parts.append(describe_secondary_pattern(p))

    return "\n".join(parts) if parts else None


def session_to_duration(session: str) -> str:
    """Map Session code → duration_type. Raises on unknown codes."""
    if session not in SESSION_TO_DURATION:
        raise ValueError(
            f"Unknown Session code {session!r}. Add it to SESSION_TO_DURATION "
            f"in fall_2026.py before re-running."
        )
    return SESSION_TO_DURATION[session]


def build_course_row(class_nbr: int, rows: list[dict[str, Any]]) -> dict[str, Any]:
    """Aggregate the source rows for one Class Nbr into a single courses row."""
    first = rows[0]
    patterns = sorted([build_pattern(r) for r in rows], key=lambda p: p["pat_nbr"])
    primary = pick_primary(patterns)

    title = (first.get("Descr") or "").strip()
    session = (first.get("Session") or "").strip()
    duration = session_to_duration(session)

    return {
        "class_nbr":          str(class_nbr),
        "subject":            first.get("Subject"),
        "catalog":            str(first.get("Catalog") or "").strip() or None,
        "course_title":       title or None,
        "course_name":        title or None,  # legacy column, populated as clean title
        "credits":            float(first["Credits"]) if first.get("Credits") is not None else None,
        "section":            str(first.get("Section") or "").strip() or None,
        "instructor":         (first.get("Name") or "").strip() or None,
        "session_code":       session,
        "duration_type":      duration,
        "mode":               (first.get("Mode") or "").strip() or None,
        "term_code":          str(first.get("Term") or "").strip() or None,

        # Primary pattern → flat columns (back-compat for search, badges, etc.)
        "meeting_days":       primary["meeting_days"] or None,
        "start_time":         primary["start_time"],
        "end_time":           primary["end_time"],
        "dates_full":         primary["dates_full"] or None,
        "meeting_times_full": format_meeting_times_full(
            primary["meeting_days"],
            time.fromisoformat(primary["start_time"]) if primary["start_time"] else None,
            time.fromisoformat(primary["end_time"]) if primary["end_time"] else None,
        ),

        # Full pattern array (always at least 1 entry).
        "meeting_patterns":   patterns,

        "notes":              build_notes(rows, primary["pat_nbr"], (first.get("Mode") or "").strip() or None),
    }


# ---- xlsx parsing -----------------------------------------------------

def load_rows(xlsx_path: Path) -> list[dict[str, Any]]:
    """Read xlsx and return a list of row-dicts keyed by header name.

    Handles the two duplicated 'Descr' headers by renaming the second to
    'Descr_2'. Drops fully-empty rows.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    raw_headers = [cell.value for cell in ws[1]]

    headers: list[str] = []
    seen: dict[str, int] = {}
    for h in raw_headers:
        name = h if h is not None else ""
        if name in seen:
            seen[name] += 1
            headers.append(f"{name}_{seen[name]}")
        else:
            seen[name] = 1
            headers.append(name)

    rows: list[dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        rows.append({h: v for h, v in zip(headers, r) if h})
    return rows


def group_by_class_nbr(rows: list[dict[str, Any]]) -> dict[int, list[dict[str, Any]]]:
    grouped: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for r in rows:
        cn = r.get("Class Nbr")
        if cn is None:
            continue
        grouped[int(cn)].append(r)
    # Sort each class's rows by Pat Nbr for stable output.
    for cn in grouped:
        grouped[cn].sort(key=lambda x: int(x.get("Pat Nbr") or 0))
    return grouped


# ---- Output modes -----------------------------------------------------

def summarize(course_rows: list[dict[str, Any]]) -> None:
    """Print row counts grouped by duration_type and Session."""
    by_dur = Counter(c["duration_type"] for c in course_rows)
    print(f"\nTotal courses: {len(course_rows)}")
    print("\nBy duration_type:")
    for k in ("Full Semester", "First Half", "Second Half", "Intensive", "DBi"):
        print(f"  {k:<14} {by_dur.get(k, 0)}")
    print(f"\nMulti-pattern classes: "
          f"{sum(1 for c in course_rows if len(c['meeting_patterns']) > 1)}")


def sql_literal(v: Any) -> str:
    """Format a Python value as a Postgres literal."""
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)):
        return repr(v)
    if isinstance(v, (dict, list)):
        # JSONB literal.
        return "'" + json.dumps(v).replace("'", "''") + "'::jsonb"
    # Treat as string: escape single quotes by doubling.
    return "'" + str(v).replace("'", "''") + "'"


def emit_sql(course_rows: list[dict[str, Any]], out_path: Path) -> None:
    """Write a single multi-row INSERT...ON CONFLICT statement (idempotent)."""
    cols = [
        "class_nbr", "subject", "catalog", "course_title", "course_name",
        "credits", "section", "instructor", "session_code", "duration_type",
        "mode", "term_code", "meeting_days", "start_time", "end_time",
        "dates_full", "meeting_times_full", "meeting_patterns", "notes",
    ]
    update_cols = [c for c in cols if c != "class_nbr"]
    update_clause = ", ".join(f"{c} = EXCLUDED.{c}" for c in update_cols)

    value_rows = [
        "(" + ", ".join(sql_literal(c[col]) for col in cols) + ")"
        for c in course_rows
    ]

    sql = (
        "-- Auto-generated by scripts/ingest/fall_2026.py — do not hand-edit.\n"
        f"INSERT INTO public.courses ({', '.join(cols)}) VALUES\n"
        + ",\n".join(value_rows)
        + f"\nON CONFLICT (class_nbr) DO UPDATE SET {update_clause};\n"
    )
    out_path.write_text(sql, encoding="utf-8")
    print(f"\nWrote multi-row INSERT for {len(course_rows)} courses to {out_path}")


def upsert_via_supabase(course_rows: list[dict[str, Any]]) -> None:
    """Direct write via Supabase PostgREST + SUPABASE_SERVICE_ROLE_KEY.

    Uses stdlib urllib to avoid the supabase-py dependency chain
    (pyiceberg fails to build on Python 3.14 as of late 2025).
    """
    import urllib.error
    import urllib.request

    try:
        from dotenv import load_dotenv  # type: ignore
    except ImportError:
        sys.exit(
            "Missing deps. From scripts/ingest/ run:\n"
            "    python -m pip install -r requirements.txt"
        )

    here = Path(__file__).parent
    load_dotenv(here / ".env")
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        sys.exit(
            "Missing env. Set SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY "
            "in scripts/ingest/.env (see README.md)."
        )

    endpoint = f"{url.rstrip('/')}/rest/v1/courses?on_conflict=class_nbr"
    payload = json.dumps(course_rows).encode("utf-8")

    req = urllib.request.Request(
        endpoint,
        data=payload,
        method="POST",
        headers={
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
            # Upsert + return only counts (avoid huge response body):
            "Prefer": "resolution=merge-duplicates,return=minimal,count=exact",
        },
    )
    try:
        with urllib.request.urlopen(req) as resp:
            count_header = resp.headers.get("Content-Range", "")
            print(f"\nUpserted (HTTP {resp.status}). Content-Range: {count_header!r}")
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        sys.exit(f"\nUpsert failed (HTTP {e.code}): {body}")


# ---- Entrypoint -------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Fall 2026 catalog ingest.")
    parser.add_argument(
        "--xlsx",
        default=str(Path.home() / "Downloads" / "Fall 2026 MBA course schedule.xlsx"),
        help="Path to the source xlsx (default: ~/Downloads/Fall 2026 MBA course schedule.xlsx).",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Parse and summarize, but don't write anywhere.",
    )
    parser.add_argument(
        "--sql-out", metavar="FILE",
        help="Emit INSERT...ON CONFLICT SQL to FILE instead of upserting directly.",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        sys.exit(f"xlsx not found: {xlsx_path}")

    print(f"Reading {xlsx_path}")
    rows = load_rows(xlsx_path)
    print(f"  {len(rows)} source rows")

    grouped = group_by_class_nbr(rows)
    print(f"  {len(grouped)} distinct Class Nbrs")

    course_rows: list[dict[str, Any]] = []
    for cn, class_rows in sorted(grouped.items()):
        try:
            course_rows.append(build_course_row(cn, class_rows))
        except Exception as e:
            sys.exit(f"Failed on Class Nbr {cn}: {e}")

    summarize(course_rows)

    if args.dry_run:
        print("\n(dry-run — no writes)")
        return

    if args.sql_out:
        emit_sql(course_rows, Path(args.sql_out))
        return

    upsert_via_supabase(course_rows)


if __name__ == "__main__":
    main()
